"""Create a simple, school-facing Excel workbook from ExamPen results."""

from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any, Iterable, Mapping, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_STATUS_LABELS = {
    "expected": "Not submitted",
    "submitted": "Submitted",
    "evaluating": "Checking",
    "blocked": "Needs attention",
    "review": "Under review",
    "ready": "Ready",
    "published": "Published",
}

_TITLE_FILL = PatternFill("solid", fgColor="0F766E")
_HEADER_FILL = PatternFill("solid", fgColor="DFF5EE")
_ALT_ROW_FILL = PatternFill("solid", fgColor="F8FAFC")
_HEADER_BORDER = Border(bottom=Side(style="thin", color="99C9BA"))


def _safe_excel_text(value: Any) -> str:
    """Prevent user-controlled names or identifiers from becoming formulas."""

    text = str(value or "")
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _filename_slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^A-Za-z0-9]+", "-", normalized).strip("-").lower()
    return slug[:80] or "exam"


def exam_marks_filename(exam_title: str) -> str:
    return f"{_filename_slug(exam_title)}-student-marks.xlsx"


def _add_sheet_heading(
    sheet: Any,
    *,
    exam_title: str,
    subtitle: str,
    final_column: str,
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(f"A1:{final_column}1")
    sheet["A1"] = _safe_excel_text(exam_title)
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = _TITLE_FILL
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(f"A2:{final_column}2")
    sheet["A2"] = _safe_excel_text(subtitle)
    sheet["A2"].font = Font(size=10, bold=True, color="475569")
    sheet["A2"].alignment = Alignment(vertical="center")


def _add_headers(sheet: Any, headers: list[str], *, row: int) -> None:
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="134E4A")
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 28


def _student_status(
    *,
    roster_status: str,
    result: Optional[Mapping[str, Any]],
) -> str:
    publication_status = str((result or {}).get("publication_status") or "").lower()
    score_state = str((result or {}).get("score_state") or "available").lower()
    if publication_status == "published":
        return "Published"
    if result and score_state == "processing":
        return "Checking"
    if result and score_state == "unavailable":
        return "Needs attention"
    return _STATUS_LABELS.get(
        roster_status,
        roster_status.replace("_", " ").title(),
    )


def build_exam_marks_workbook(
    *,
    exam_title: str,
    class_label: Optional[str],
    roster_rows: Iterable[Mapping[str, Any]],
    result_rows: Iterable[Mapping[str, Any]],
    question_rows: Iterable[Mapping[str, Any]],
) -> bytes:
    """Return a school-friendly workbook with marks and question accuracy."""

    roster = list(roster_rows)
    questions = list(question_rows)
    results_by_student = {
        str(row.get("student_id") or ""): row
        for row in result_rows
        if str(row.get("student_id") or "")
    }

    workbook = Workbook()
    workbook.properties.title = f"{exam_title} student marks"
    workbook.properties.subject = "Student marks and question-wise class accuracy"
    workbook.properties.creator = "Stoody"

    marks_sheet = workbook.active
    marks_sheet.title = "Student Marks"
    marks_headers = [
        "S.No.",
        "Student name",
        "Student ID",
        "Marks scored",
        "Total marks",
        "Percentage",
        "Status",
    ]
    marks_header_row = 3
    _add_sheet_heading(
        marks_sheet,
        exam_title=exam_title,
        subtitle=(f"Student Marks | {class_label}" if class_label else "Student Marks"),
        final_column="G",
    )
    _add_headers(marks_sheet, marks_headers, row=marks_header_row)

    for sequence, roster_row in enumerate(roster, start=1):
        student_id = str(roster_row.get("student_id") or "")
        result = results_by_student.get(student_id)
        roster_status = str(roster_row.get("status") or "expected").lower()
        publication_status = str((result or {}).get("publication_status") or "").lower()
        score_state = str((result or {}).get("score_state") or "available").lower()
        score_available = bool(
            result
            and (publication_status == "published" or score_state == "available")
            and float((result or {}).get("combined_max") or 0) > 0
        )
        scored = (
            float(result.get("combined_total") or 0)
            if score_available and result
            else None
        )
        maximum = (
            float(result.get("combined_max") or 0)
            if score_available and result
            else None
        )
        values: list[Any] = [
            sequence,
            _safe_excel_text(roster_row.get("student_name") or student_id),
            _safe_excel_text(student_id),
            scored,
            maximum,
            scored / maximum if scored is not None and maximum else None,
            _student_status(roster_status=roster_status, result=result),
        ]
        row_number = marks_header_row + sequence
        for column, value in enumerate(values, start=1):
            cell = marks_sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if column in {1, 4, 5, 6} else "left",
            )
            if row_number % 2 == 0:
                cell.fill = _ALT_ROW_FILL
        marks_sheet.cell(row=row_number, column=4).number_format = "0.##"
        marks_sheet.cell(row=row_number, column=5).number_format = "0.##"
        marks_sheet.cell(row=row_number, column=6).number_format = "0.0%"

    marks_last_row = marks_header_row + len(roster)
    marks_sheet.freeze_panes = f"A{marks_header_row + 1}"
    marks_sheet.auto_filter.ref = (
        f"A{marks_header_row}:G{max(marks_last_row, marks_header_row)}"
    )
    marks_sheet.print_title_rows = f"1:{marks_header_row}"
    for column, width in {
        1: 8,
        2: 26,
        3: 22,
        4: 16,
        5: 14,
        6: 13,
        7: 18,
    }.items():
        marks_sheet.column_dimensions[get_column_letter(column)].width = width

    accuracy_sheet = workbook.create_sheet("Question Accuracy")
    accuracy_headers = [
        "Question No.",
        "Question",
        "Maximum marks",
        "Students assessed",
        "Average marks",
        "Class accuracy",
    ]
    accuracy_header_row = 4
    _add_sheet_heading(
        accuracy_sheet,
        exam_title=exam_title,
        subtitle="Question-wise Class Accuracy",
        final_column="F",
    )
    accuracy_sheet.merge_cells("A3:F3")
    accuracy_sheet["A3"] = (
        "Class accuracy = total marks scored by assessed students "
        "÷ total possible marks for those students."
    )
    accuracy_sheet["A3"].font = Font(size=9, italic=True, color="64748B")
    accuracy_sheet["A3"].alignment = Alignment(vertical="center")
    _add_headers(accuracy_sheet, accuracy_headers, row=accuracy_header_row)

    for sequence, question in enumerate(questions, start=1):
        assessed_count = int(question.get("assessed_count") or 0)
        question_number = question.get("question_number") or sequence
        average_marks = (
            float(question.get("average_score") or 0) if assessed_count > 0 else None
        )
        accuracy = (
            float(question.get("average_percent") or 0) / 100
            if assessed_count > 0
            else None
        )
        values = [
            question_number,
            _safe_excel_text(
                question.get("question_text") or f"Question {question_number}"
            ),
            float(question.get("max_marks") or 0),
            assessed_count,
            average_marks,
            accuracy,
        ]
        row_number = accuracy_header_row + sequence
        for column, value in enumerate(values, start=1):
            cell = accuracy_sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(
                vertical="top",
                horizontal="right" if column in {1, 3, 4, 5, 6} else "left",
                wrap_text=column == 2,
            )
            if row_number % 2 == 0:
                cell.fill = _ALT_ROW_FILL
        accuracy_sheet.cell(row=row_number, column=3).number_format = "0.##"
        accuracy_sheet.cell(row=row_number, column=5).number_format = "0.##"
        accuracy_sheet.cell(row=row_number, column=6).number_format = "0.0%"

    if not questions:
        accuracy_sheet.merge_cells("A5:F5")
        accuracy_sheet["A5"] = "Question-wise accuracy is not available yet."
        accuracy_sheet["A5"].font = Font(italic=True, color="64748B")
        accuracy_sheet["A5"].alignment = Alignment(horizontal="center")

    accuracy_last_row = accuracy_header_row + max(len(questions), 1)
    accuracy_sheet.freeze_panes = f"A{accuracy_header_row + 1}"
    accuracy_sheet.auto_filter.ref = (
        f"A{accuracy_header_row}:F{accuracy_last_row}"
        if questions
        else f"A{accuracy_header_row}:F{accuracy_header_row}"
    )
    accuracy_sheet.print_title_rows = f"1:{accuracy_header_row}"
    for column, width in {
        1: 14,
        2: 60,
        3: 16,
        4: 18,
        5: 16,
        6: 16,
    }.items():
        accuracy_sheet.column_dimensions[get_column_letter(column)].width = width

    for sheet in (marks_sheet, accuracy_sheet):
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
