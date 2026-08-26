from io import BytesIO
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook
import pytest


def test_exam_marks_workbook_is_simple_and_hides_stale_scores():
    from services.exam_marks_export import build_exam_marks_workbook

    content = build_exam_marks_workbook(
        exam_title="Vectors and Projectile",
        class_label="Class 10 - Section A",
        roster_rows=[
            {
                "student_id": "aaradhya",
                "student_name": "Aaradhya Tomar",
                "status": "published",
            },
            {
                "student_id": "missing-student",
                "student_name": "Missing Student",
                "status": "expected",
            },
            {
                "student_id": "stale-student",
                "student_name": '=HYPERLINK("https://invalid.example")',
                "status": "blocked",
            },
        ],
        result_rows=[
            {
                "student_id": "aaradhya",
                "combined_total": 17,
                "combined_max": 25,
                "publication_status": "published",
                "score_state": "available",
            },
            {
                "student_id": "stale-student",
                "combined_total": 14,
                "combined_max": 25,
                "publication_status": "pending",
                "score_state": "unavailable",
            },
        ],
        question_rows=[
            {
                "question_number": 1,
                "question_text": "Find x.",
                "max_marks": 5,
                "assessed_count": 2,
                "average_score": 3.5,
                "average_percent": 70,
            },
            {
                "question_number": 2,
                "question_text": "Draw the vector.",
                "max_marks": 4,
                "assessed_count": 0,
                "average_score": 0,
                "average_percent": 0,
            },
        ],
    )

    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["Student Marks", "Question Accuracy"]

    marks = workbook["Student Marks"]
    assert marks["A1"].value == "Vectors and Projectile"
    assert marks["A2"].value == "Student Marks | Class 10 - Section A"
    assert [cell.value for cell in marks[3]] == [
        "S.No.",
        "Student name",
        "Student ID",
        "Marks scored",
        "Total marks",
        "Percentage",
        "Status",
    ]
    assert marks.freeze_panes == "A4"
    assert marks.auto_filter.ref == "A3:G6"

    assert marks["B4"].value == "Aaradhya Tomar"
    assert marks["D4"].value == 17
    assert marks["E4"].value == 25
    assert marks["F4"].value == 17 / 25
    assert marks["G4"].value == "Published"

    assert marks["D5"].value is None
    assert marks["E5"].value is None
    assert marks["F5"].value is None
    assert marks["G5"].value == "Not submitted"

    assert marks["B6"].value.startswith("'=")
    assert marks["D6"].value is None
    assert marks["E6"].value is None
    assert marks["F6"].value is None
    assert marks["G6"].value == "Needs attention"

    accuracy = workbook["Question Accuracy"]
    assert [cell.value for cell in accuracy[4]] == [
        "Question No.",
        "Question",
        "Maximum marks",
        "Students assessed",
        "Average marks",
        "Class accuracy",
    ]
    assert accuracy.freeze_panes == "A5"
    assert accuracy.auto_filter.ref == "A4:F6"
    assert accuracy["A5"].value == 1
    assert accuracy["B5"].value == "Find x."
    assert accuracy["C5"].value == 5
    assert accuracy["D5"].value == 2
    assert accuracy["E5"].value == 3.5
    assert accuracy["F5"].value == 0.7
    assert accuracy["E6"].value is None
    assert accuracy["F6"].value is None

    visible_text = " ".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ).lower()
    assert "pcr" not in visible_text
    assert "dcr" not in visible_text
    assert "recheck" not in visible_text
    assert "exam id" not in visible_text


def test_exam_marks_filename_is_safe_and_readable():
    from services.exam_marks_export import exam_marks_filename

    assert (
        exam_marks_filename("Vectors & Projectile / Class 10")
        == "vectors-projectile-class-10-student-marks.xlsx"
    )


@pytest.mark.asyncio
async def test_export_endpoint_returns_marks_and_persisted_question_accuracy():
    from mongomock_motor import AsyncMongoMockClient

    from api.v1.evalpen_review_async import (
        CollectionRosterItemAPI,
        ExamResultStudentAPI,
        ExamResultsAPI,
        ExamRosterAPI,
        export_exam_results,
    )

    tenant_db = AsyncMongoMockClient()["skb_test"]
    await tenant_db["exampen_exams"].insert_one(
        {
            "exam_id": "EXAM-EXPORT",
            "title": "Vectors and Projectile",
            "class_name": "10",
            "section_name": "A",
        }
    )
    roster = ExamRosterAPI(
        exam_id="EXAM-EXPORT",
        expected_students=[
            CollectionRosterItemAPI(
                student_id="student-1",
                student_name="Student One",
                status="published",
            )
        ],
        total_expected=1,
        total_submitted=1,
        total_published=1,
    )
    results = ExamResultsAPI(
        exam_id="EXAM-EXPORT",
        students=[
            ExamResultStudentAPI(
                student_id="student-1",
                combined_total=22,
                combined_max=25,
                pcr_total_score=22,
                pcr_max_score=25,
                publication_status="published",
            )
        ],
        total_students=1,
    )
    analytics = {
        "questions": [
            {
                "question_id": "question-1",
                "question_number": 1,
                "question_text": "Find x.",
                "max_marks": 5,
                "assessed_count": 0,
                "average_score": 0,
                "average_percent": 0,
            }
        ]
    }
    await tenant_db["exampen_dcr_results"].insert_one(
        {
            "exam_id": "EXAM-EXPORT",
            "student_id": "student-1",
            "question_id": "question-1",
            "score": 4,
            "max_score": 5,
        }
    )

    with (
        patch(
            "api.v1.evalpen_review_async.get_exam_roster",
            new=AsyncMock(return_value=roster),
        ),
        patch(
            "api.v1.evalpen_review_async.get_exam_results",
            new=AsyncMock(return_value=results),
        ),
        patch(
            "api.v1.evalpen_review_async.get_exam_analytics",
            new=AsyncMock(return_value=analytics),
        ),
        patch(
            "api.v1.evalpen_review_async._get_tenant_db",
            new=AsyncMock(return_value=tenant_db),
        ),
    ):
        response = await export_exam_results(
            exam_id="EXAM-EXPORT",
            current_user={
                "user_id": "admin-1",
                "user_type": "admin",
                "db_name": "skb_test",
            },
            db=None,
        )

    content = b"".join([chunk async for chunk in response.body_iterator])
    workbook = load_workbook(BytesIO(content))
    marks = workbook["Student Marks"]
    accuracy = workbook["Question Accuracy"]

    assert response.media_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="vectors-and-projectile-student-marks.xlsx"'
    )
    assert marks["B4"].value == "Student One"
    assert marks["D4"].value == 22
    assert marks["E4"].value == 25
    assert accuracy["B5"].value == "Find x."
    assert accuracy["F5"].value == 0.8
